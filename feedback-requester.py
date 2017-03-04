# Feedback Request Sender

import mechanize
import openpyxl

# avoid unicode error
import sys
reload(sys)
sys.setdefaultencoding('utf-8')


'''
# Use mechanize to find correct form to fill out

br = mechanize.Browser()
br.open("<property webpage>")

for f in br.forms():
	print f

# the above shows that there is no name for the form, so we find it by index below

br.form=list(br.forms())[2]
print br.form

# View form control types, names, values

for control in br.form.controls:
	print control
	print "type=%s, name=%s, value=%s" % (control.type, control.name, br[control.name])

'''

# put clients in excel file by copy pasting into a new column each contact:

# name      name      name
# email     email     email
# phone     phone     phone
# company   company   company

# Get contact info from excel file by assigning each field to a list in contact order

wb = openpyxl.load_workbook(filename='test.xlsx')
ws = wb['Sheet2']

name = []
email = []
phone = []
company = []


for cell in ws['1']:
	name.append(cell.value)
for cell in ws['2']:
	email.append(cell.value)
for cell in ws['3']:
	phone.append(cell.value)
for cell in ws['4']:
	company.append(cell.value)

col_count = ws.max_column

# For each contact


print "filling out form for <property address>..." ####### type in address form is being sent for, and change url below

for contact in range(0,col_count):

	# get form from webpage

	br = mechanize.Browser()
	br.open("<property webpage>") ####### property page url 
	br.form=list(br.forms())[2]  # the correct form


	# fill out the form by assigning contact's info from excel list to form field, then submit form

	br['Name'] = str(name[contact]) 
	br['Company'] = str(company[contact])
	br['Phone'] = str(phone[contact])
	br['Email'] = str(email[contact])

	response = br.submit()

	print "sent request to " + name[contact]
	br.back() # I think this makes it look human?? idk


print "done!"
